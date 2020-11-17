import openpyxl

# Abro el archivo/libro de excel
mixls = openpyxl.load_workbook("nuevoExcel.xlsx")

# Crear nueva hoja
mixls.create_sheet()
# Imprimo las hojas del libro de excel
print(mixls.sheetnames)

# Creo nueva hoja en la primera posición
mixls.create_sheet(index=0, title="1ra hoja")
# Imprimo las hojas
print(mixls.sheetnames)

# Creo otra hoja en la posición 2 que es la 3ra hoja
mixls.create_sheet(index=2, title="3ra hoja")
# Imprimo las hojas
print(mixls.sheetnames)

# Borro una hoja de excel
mixls.remove( mixls["1ra hoja"] )
# Tambien podemos usar esta instrucción
# del mixls["1ra hoja"]

# Seleciono una hoja de Excel
hoja = mixls["3ra hoja"]

# Asigno el valor de la celda F6
hoja["F6"] = "Valor de F6"

#obtengo el valor de la celda F6
print(hoja["F6"].value)

# Guardo los cambios
mixls.save("otroExcel.xlsx")

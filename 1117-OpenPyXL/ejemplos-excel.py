import openpyxl

# Creamos un nuevo archivo/Libro de Excel
mixls = openpyxl.Workbook()

# Imprimo los nombres de las hojas
print(mixls.sheetnames)

# Selecciono la hoja activa
hoja = mixls.active

# Imprimimos el titulo de la hoja
print("Titulo de la hoja activa:",hoja.title)

# Modificamos el titulo de la hoja
hoja.title = "Nuevo nombre"

# Imprimimos las hojas del libro
print(mixls.sheetnames)

# Guardamos este archivo de excel con el nombre nuevoExcel.xlsx
mixls.save("nuevoExcel.xlsx")



